"""
exportar.py
Exportación del calendario a Excel y PDF con formato profesional.
"""

import pandas as pd
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

# ------------------------------------------------------------
# Constantes de colores (sin '#', para compatibilidad con openpyxl)
# ------------------------------------------------------------
COLOR_MORNING = "00B050"      # Verde
COLOR_AFTERNOON = "FFC000"    # Amarillo/Naranja
COLOR_NIGHT = "A6A6A6"        # Gris
COLOR_FESTIVO = "FF0000"      # Rojo
COLOR_TEXT_DARK = "000000"    # Negro
COLOR_TEXT_LIGHT = "FFFFFF"   # Blanco

# ------------------------------------------------------------
# Constantes para PDF (con '#')
# ------------------------------------------------------------
PDF_COLOR_MORNING = "27AE60"
PDF_COLOR_AFTERNOON = "F39C12"
PDF_COLOR_NIGHT = "34495E"
PDF_COLOR_HOLIDAY = "E74C3C"
PDF_COLOR_TEXT = "2C3E50"

def _dia_abreviatura(fecha: datetime) -> str:
    """Retorna la abreviatura de una letra para el día de la semana."""
    dias = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
    return dias[fecha.weekday()]

# ------------------------------------------------------------
# PDF
# ------------------------------------------------------------
def generar_pdf_calendario(df: pd.DataFrame, año: int, mes: int, ciclo: str, output_path: str) -> None:
    """Genera un PDF con el calendario de turnos."""
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)

    elementos = []
    styles = getSampleStyleSheet()
    titulo_style = styles['Title']
    titulo_style.textColor = colors.HexColor("#" + PDF_COLOR_TEXT)

    titulo = Paragraph(f"PROGRAMACIÓN MES DE {calendar.month_name[mes].upper()} {año} (Ciclo {ciclo})", titulo_style)
    elementos.append(titulo)
    elementos.append(Spacer(1, 0.5*cm))

    # Preparar datos
    personas = df[['persona_id', 'nombre']].drop_duplicates().sort_values('persona_id')
    fechas = sorted(df['fecha'].unique())
    fechas_obj = [datetime.strptime(f, "%Y-%m-%d").date() for f in fechas]

    # Construir tabla
    data = []
    # Fila 1: día de la semana
    fila1 = [""] + [_dia_abreviatura(f) for f in fechas_obj]
    data.append(fila1)
    # Fila 2: número del día
    fila2 = ["NOMBRE"] + [str(f.day) for f in fechas_obj]
    data.append(fila2)

    for _, persona in personas.iterrows():
        pid = persona['persona_id']
        nombre = persona['nombre']
        fila = [nombre]
        for fecha in fechas:
            turno_df = df[(df['persona_id'] == pid) & (df['fecha'] == fecha)]
            fila.append(turno_df['turno'].values[0] if not turno_df.empty else "")
        data.append(fila)

    ancho_columna = 1.2 * cm
    col_widths = [3*cm] + [ancho_columna] * len(fechas)
    tabla = Table(data, colWidths=col_widths)

    style = []

    # Fondo de la primera columna (nombres)
    for i in range(2, len(data)):
        style.append(('BACKGROUND', (0,i), (0,i), colors.HexColor("#" + PDF_COLOR_TEXT)))
        style.append(('TEXTCOLOR', (0,i), (0,i), colors.white))

    # Cabeceras
    style.append(('BACKGROUND', (0,0), (-1,0), colors.HexColor("#" + PDF_COLOR_TEXT)))
    style.append(('TEXTCOLOR', (0,0), (-1,0), colors.white))
    style.append(('BACKGROUND', (0,1), (-1,1), colors.HexColor("#" + PDF_COLOR_TEXT)))
    style.append(('TEXTCOLOR', (0,1), (-1,1), colors.white))

    style.append(('ALIGN', (0,0), (-1,-1), 'CENTER'))
    style.append(('VALIGN', (0,0), (-1,-1), 'MIDDLE'))
    style.append(('FONTNAME', (0,0), (-1,1), 'Helvetica-Bold'))
    style.append(('FONTSIZE', (0,0), (-1,-1), 8))
    style.append(('GRID', (0,0), (-1,-1), 0.5, colors.grey))

    # Colores por turno en celdas de datos
    for i, fila in enumerate(data[2:], start=2):
        for j, fecha in enumerate(fechas, start=1):
            turno = data[i][j]
            if turno in ['V', 'PNR', 'IM', 'CD']:
                style.append(('BACKGROUND', (j,i), (j,i), colors.HexColor("#" + PDF_COLOR_HOLIDAY)))
                style.append(('TEXTCOLOR', (j,i), (j,i), colors.white))
            elif turno == 'M':
                style.append(('BACKGROUND', (j,i), (j,i), colors.HexColor("#" + PDF_COLOR_MORNING)))
                style.append(('TEXTCOLOR', (j,i), (j,i), colors.white))
            elif turno == 'T':
                style.append(('BACKGROUND', (j,i), (j,i), colors.HexColor("#" + PDF_COLOR_AFTERNOON)))
                style.append(('TEXTCOLOR', (j,i), (j,i), colors.white))
            elif turno == 'N':
                style.append(('BACKGROUND', (j,i), (j,i), colors.HexColor("#" + PDF_COLOR_NIGHT)))
                style.append(('TEXTCOLOR', (j,i), (j,i), colors.white))

    # Resaltar días festivos en cabecera
    for j, fecha in enumerate(fechas, start=1):
        es_festivo = df[df['fecha'] == fecha]['es_festivo'].max()
        if es_festivo:
            style.append(('BACKGROUND', (j,0), (j,0), colors.HexColor("#" + PDF_COLOR_HOLIDAY)))
            style.append(('BACKGROUND', (j,1), (j,1), colors.HexColor("#" + PDF_COLOR_HOLIDAY)))
            style.append(('TEXTCOLOR', (j,0), (j,1), colors.white))

    tabla.setStyle(TableStyle(style))
    elementos.append(tabla)
    elementos.append(Spacer(1, 0.5*cm))

    leyenda_style = ParagraphStyle('Leyenda', parent=styles['Normal'], fontSize=8)
    leyenda = Paragraph(
        "🟢 M = Mañana   🟡 T = Tarde   🔵 N = Noche   🔴 Festivo/Domingo/Ausencia",
        leyenda_style
    )
    elementos.append(leyenda)

    doc.build(elementos)

# ------------------------------------------------------------
# Excel (formato profesional con imagen)
# ------------------------------------------------------------
def generar_excel_con_formato(df: pd.DataFrame, año: int, mes: int, ciclo: str, output_path: str) -> None:
    """
    Genera un archivo Excel con el formato profesional exacto de la imagen adjunta.
    """
    # Preparar datos
    personas = df[['persona_id', 'nombre']].drop_duplicates().sort_values('persona_id')
    fechas = sorted(df['fecha'].unique())
    fechas_obj = [datetime.strptime(f, "%Y-%m-%d").date() for f in fechas]
    n_dias = len(fechas)
    
    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = f"Calendario {mes}-{año}"
    
    # --- Estilos básicos ---
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Encabezado institucional ---
    # Fila 1: AÑO y título
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=n_dias+1)
    celda_titulo = ws.cell(row=1, column=3)
    celda_titulo.value = "Tennis Scheduling Solutions"
    celda_titulo.font = Font(size=14, bold=True)
    celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Celda A1: "AÑO"
    ws.cell(row=1, column=1, value="AÑO").font = Font(bold=True)
    ws.cell(row=1, column=1).border = thin_border
    # Celda B1: año
    ws.cell(row=1, column=2, value=año).border = thin_border
    
    # Fila 2: MES
    ws.cell(row=2, column=1, value="MES").font = Font(bold=True)
    ws.cell(row=2, column=1).border = thin_border
    ws.cell(row=2, column=2, value=calendar.month_name[mes]).border = thin_border
    
    # --- Cabecera de días ---
    # Fila 3: iniciales de los días
    ws.cell(row=3, column=1, value="DIAS").font = Font(bold=True)
    ws.cell(row=3, column=1).border = thin_border
    # Fila 4: números de día
    ws.cell(row=4, column=1, value="NOMBRE").font = Font(bold=True)
    ws.cell(row=4, column=1).border = thin_border
    
    # Rellenar columnas de días
    for i, fecha in enumerate(fechas_obj, start=2):
        col = i
        # Día de la semana (inicial)
        ws.cell(row=3, column=col, value=_dia_abreviatura(fecha))
        # Número del día
        ws.cell(row=4, column=col, value=fecha.day)
        # Estilo de cabecera: fondo rojo si es domingo o festivo
        es_festivo = df[df['fecha'] == fecha.isoformat()]['es_festivo'].max() if not df[df['fecha'] == fecha.isoformat()].empty else 0
        if fecha.weekday() == 6 or es_festivo:
            fill_red = PatternFill(start_color=COLOR_FESTIVO, end_color=COLOR_FESTIVO, fill_type="solid")
            font_negro = Font(color=COLOR_TEXT_DARK, bold=True)
            for fila in [3, 4]:
                celda = ws.cell(row=fila, column=col)
                celda.fill = fill_red
                celda.font = font_negro
        else:
            # Fondo gris claro para días normales
            fill_gris = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            for fila in [3, 4]:
                celda = ws.cell(row=fila, column=col)
                celda.fill = fill_gris
                celda.font = Font(bold=True)
        # Bordes para las celdas de cabecera
        for fila in [3, 4]:
            ws.cell(row=fila, column=col).border = thin_border
    
    # --- Datos de personas ---
    start_row = 5
    for idx, (_, persona) in enumerate(personas.iterrows(), start=start_row):
        pid = persona['persona_id']
        nombre = persona['nombre']
        ws.cell(row=idx, column=1, value=nombre).border = thin_border
        # Fondo de la primera columna (nombre) oscuro
        ws.cell(row=idx, column=1).fill = PatternFill(start_color=COLOR_NIGHT, end_color=COLOR_NIGHT, fill_type="solid")
        ws.cell(row=idx, column=1).font = Font(color=COLOR_TEXT_LIGHT, bold=True)
        ws.cell(row=idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        for col, fecha in enumerate(fechas, start=2):
            turno_df = df[(df['persona_id'] == pid) & (df['fecha'] == fecha)]
            turno = turno_df['turno'].values[0] if not turno_df.empty else ""
            celda = ws.cell(row=idx, column=col, value=turno)
            celda.border = thin_border
            celda.alignment = Alignment(horizontal='center', vertical='center')
            
            # Determinar colores
            es_festivo = turno_df['es_festivo'].max() if not turno_df.empty else 0
            es_domingo = datetime.strptime(fecha, "%Y-%m-%d").date().weekday() == 6
            
            if turno in ['V', 'PNR', 'IM', 'CD']:
                # Ausencia: rojo con texto negro
                celda.fill = PatternFill(start_color=COLOR_FESTIVO, end_color=COLOR_FESTIVO, fill_type="solid")
                celda.font = Font(color=COLOR_TEXT_DARK)
            elif es_festivo or es_domingo:
                # Festivo o domingo: rojo con texto negro
                celda.fill = PatternFill(start_color=COLOR_FESTIVO, end_color=COLOR_FESTIVO, fill_type="solid")
                celda.font = Font(color=COLOR_TEXT_DARK)
            else:
                # Día normal: colorear según turno
                if turno == 'M':
                    celda.fill = PatternFill(start_color=COLOR_MORNING, end_color=COLOR_MORNING, fill_type="solid")
                    celda.font = Font(color=COLOR_TEXT_LIGHT)
                elif turno == 'T':
                    celda.fill = PatternFill(start_color=COLOR_AFTERNOON, end_color=COLOR_AFTERNOON, fill_type="solid")
                    celda.font = Font(color=COLOR_TEXT_LIGHT)
                elif turno == 'N':
                    celda.fill = PatternFill(start_color=COLOR_NIGHT, end_color=COLOR_NIGHT, fill_type="solid")
                    celda.font = Font(color=COLOR_TEXT_LIGHT)
                else:
                    # Si es L o cualquier otro, sin fondo
                    celda.fill = PatternFill(fill_type=None)
                    celda.font = Font(color=COLOR_TEXT_DARK)
    
    # --- Leyenda (dos filas después de los datos) ---
    last_data_row = start_row + len(personas) - 1
    leyenda_start = last_data_row + 3  # dos filas vacías después de datos
    
    leyenda_items = [
        "M = Mañana",
        "T = Tarde",
        "N = Noche",
        "Festivo/Domingo",
        "V/PNR/IM/CD = Ausencia"
    ]
    for i, texto in enumerate(leyenda_items):
        fila = leyenda_start + i
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = Font(size=10)
        # Opcional: agregar un pequeño cuadrado de color a la izquierda (no requerido por la imagen)
        # Se deja solo el texto.
    
    # --- Ajustes de ancho de columnas ---
    # Ancho de la primera columna (nombre)
    ws.column_dimensions['A'].width = 20
    # Ancho de las columnas de días: compacto, cuadrado
    for i in range(2, n_dias + 2):
        ws.column_dimensions[get_column_letter(i)].width = 5
    
    # Guardar
    wb.save(output_path)